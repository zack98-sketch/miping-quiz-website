import re
import openpyxl
from sqlalchemy.orm import Session
from app.models import Question, Option, QuestionType, Difficulty


def import_excel(file_path: str, db: Session) -> dict:
    """Import questions from Excel file into database."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    
    success_count = 0
    error_count = 0
    errors = []
    
    # Clear existing questions (full reimport)
    db.query(Option).delete()
    db.query(Question).delete()
    db.commit()
    
    rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    total = len(rows)
    
    batch_questions = []
    batch_options = []
    
    for i, row in enumerate(rows):
        try:
            source_id = row[0]  # 序号
            raw_type = str(row[1]).strip() if row[1] else ""
            content = str(row[2]).strip() if row[2] else ""
            opt_a = str(row[3]).strip() if row[3] else ""
            opt_b = str(row[4]).strip() if row[4] else ""
            opt_c = str(row[5]).strip() if row[5] else ""
            opt_d = str(row[6]).strip() if row[6] else ""
            answer = str(row[7]).strip() if row[7] else ""
            explanation = str(row[8]).strip() if row[8] else ""
            
            # Validate
            if not content or not answer:
                error_count += 1
                errors.append({"row": i + 2, "error": "题目内容或答案为空"})
                continue
            
            # Parse question type
            question_type = _parse_question_type(raw_type)
            if not question_type:
                error_count += 1
                errors.append({"row": i + 2, "error": f"无法识别题型: {raw_type}"})
                continue
            
            # Normalize answer
            answer = _normalize_answer(answer, question_type)
            
            # Extract knowledge point from content
            knowledge_point = _extract_knowledge_point(content)
            
            # Determine difficulty (default medium)
            difficulty = _estimate_difficulty(content, explanation)
            
            question = Question(
                content=content,
                question_type=question_type,
                difficulty=difficulty,
                knowledge_point=knowledge_point,
                correct_answer=answer,
                explanation=explanation,
                source_id=int(source_id) if source_id else None,
                is_active=True,
            )
            db.add(question)
            db.flush()  # Get the ID
            
            # Add options
            if question_type == QuestionType.judge:
                options = [
                    Option(question_id=question.id, label="A", content=opt_a or "正确", sort_order=1),
                    Option(question_id=question.id, label="B", content=opt_b or "错误", sort_order=2),
                ]
            else:
                options = []
                if opt_a:
                    options.append(Option(question_id=question.id, label="A", content=opt_a, sort_order=1))
                if opt_b:
                    options.append(Option(question_id=question.id, label="B", content=opt_b, sort_order=2))
                if opt_c:
                    options.append(Option(question_id=question.id, label="C", content=opt_c, sort_order=3))
                if opt_d:
                    options.append(Option(question_id=question.id, label="D", content=opt_d, sort_order=4))
            
            for opt in options:
                db.add(opt)
            
            success_count += 1
            
            # Batch commit every 500
            if success_count % 500 == 0:
                db.commit()
                
        except Exception as e:
            error_count += 1
            errors.append({"row": i + 2, "error": str(e)})
            continue
    
    db.commit()
    wb.close()
    
    return {
        "total": total,
        "success": success_count,
        "errors": error_count,
        "error_details": errors[:50],  # Limit error details
    }


def _parse_question_type(raw_type: str) -> QuestionType | None:
    raw = raw_type.replace("∗", "").strip()
    if "单" in raw:
        return QuestionType.single
    elif "多" in raw:
        return QuestionType.multi
    elif "判断" in raw:
        return QuestionType.judge
    return None


def _normalize_answer(answer: str, question_type: QuestionType) -> str:
    answer = answer.strip().upper()
    if question_type == QuestionType.judge:
        if answer in ("正确", "对", "TRUE", "T", "1"):
            return "A"
        elif answer in ("错误", "错", "FALSE", "F", "0"):
            return "B"
    return answer


def _extract_knowledge_point(content: str) -> str:
    """Extract knowledge point from question content using keyword matching."""
    keywords_map = {
        "密码": "密码学基础",
        "加密": "加密算法",
        "解密": "加密算法",
        "对称": "对称加密",
        "非对称": "非对称加密",
        "公钥": "公钥密码",
        "私钥": "公钥密码",
        "RSA": "RSA算法",
        "AES": "AES算法",
        "DES": "DES算法",
        "哈希": "哈希算法",
        "散列": "哈希算法",
        "数字签名": "数字签名",
        "证书": "数字证书",
        "PKI": "PKI体系",
        "认证": "身份认证",
        "鉴别": "身份认证",
        "访问控制": "访问控制",
        "防火墙": "防火墙",
        "入侵": "入侵检测",
        "漏洞": "漏洞管理",
        "安全协议": "安全协议",
        "SSL": "SSL/TLS",
        "TLS": "SSL/TLS",
        "IPSec": "IPSec",
        "网络": "网络安全",
        "数据": "数据安全",
        "隐私": "隐私保护",
        "等保": "等级保护",
        "密评": "密评",
        "商密": "商用密码",
        "国密": "国产密码",
        "SM2": "国密SM2",
        "SM3": "国密SM3",
        "SM4": "国密SM4",
        "宪法": "宪法",
        "法律": "法律法规",
        "法规": "法律法规",
        "标准": "标准规范",
        "管理": "安全管理",
        "风险": "风险管理",
        "审计": "安全审计",
        "二十届": "时政",
        "全会": "时政",
        "十五五": "时政",
    }
    
    for keyword, kp in keywords_map.items():
        if keyword in content:
            return kp
    
    return "综合"


def _estimate_difficulty(content: str, explanation: str) -> Difficulty:
    """Estimate difficulty based on content length and complexity."""
    content_len = len(content)
    explanation_len = len(explanation) if explanation else 0
    
    if content_len > 200 or explanation_len > 500:
        return Difficulty.hard
    elif content_len > 100 or explanation_len > 200:
        return Difficulty.medium
    else:
        return Difficulty.easy
